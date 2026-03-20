from __future__ import annotations

import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.db.models import (
    Assembly,
    Attachment,
    AuditLog,
    ExportJob,
    MaterialRecord,
    NDTRecord,
    Project,
    Weld,
    WeldDefect,
    WeldInspection,
    WelderProfile,
    WPSRecord,
    WPQRRecord,
)
from app.services.phase5_audit import log_phase5_event

BASE_DIR = Path(__file__).resolve().parents[2]
EXPORT_ROOT = BASE_DIR / "storage" / "exports"


def _dt(value):
    if not value:
        return None
    try:
        return value.isoformat()
    except Exception:
        return str(value)


def _uuid(value):
    return str(value) if value is not None else None


def _project_or_404(db: Session, tenant_id: UUID, project_id: UUID) -> Project:
    row = db.query(Project).filter(Project.id == project_id, Project.tenant_id == tenant_id).first()
    if not row:
        raise ValueError("Project not found")
    return row


def build_preview(db: Session, tenant_id: UUID, project_id: UUID) -> dict[str, Any]:
    project = _project_or_404(db, tenant_id, project_id)

    assemblies = (
        db.query(Assembly)
        .filter(Assembly.project_id == project_id, Assembly.tenant_id == tenant_id)
        .order_by(Assembly.code.asc())
        .all()
    )
    welds = (
        db.query(Weld)
        .filter(Weld.project_id == project_id, Weld.tenant_id == tenant_id)
        .order_by(Weld.weld_no.asc())
        .all()
    )
    inspections = (
        db.query(WeldInspection)
        .filter(WeldInspection.project_id == project_id, WeldInspection.tenant_id == tenant_id)
        .all()
    )
    defects = (
        db.query(WeldDefect)
        .filter(
            WeldDefect.project_id == project_id,
            WeldDefect.tenant_id == tenant_id,
            WeldDefect.deleted_at.is_(None),
        )
        .all()
    )
    ndt_records = (
        db.query(NDTRecord)
        .filter(NDTRecord.project_id == project_id, NDTRecord.tenant_id == tenant_id)
        .all()
    )
    materials = (
        db.query(MaterialRecord)
        .filter(MaterialRecord.project_id == project_id, MaterialRecord.tenant_id == tenant_id)
        .all()
    )
    attachments = (
        db.query(Attachment)
        .filter(
            Attachment.tenant_id == tenant_id,
            Attachment.scope_type.in_(["project", "weld", "inspection"]),
        )
        .all()
    )
    audit_rows = (
        db.query(AuditLog)
        .filter(AuditLog.tenant_id == tenant_id)
        .order_by(AuditLog.created_at.desc())
        .limit(250)
        .all()
    )
    welder_profiles = (
        db.query(WelderProfile)
        .filter(WelderProfile.tenant_id == tenant_id, WelderProfile.is_active.is_(True))
        .all()
    )
    wps_records = (
        db.query(WPSRecord)
        .filter(WPSRecord.tenant_id == tenant_id, WPSRecord.is_active.is_(True))
        .all()
    )
    wpqr_records = db.query(WPQRRecord).filter(WPQRRecord.tenant_id == tenant_id).all()

    project_attachment_count = sum(
        1 for a in attachments if _uuid(a.scope_id) == _uuid(project_id) or a.scope_type in ("weld", "inspection")
    )
    accepted_results = sum(1 for r in inspections if (r.overall_status or "") == "ok")
    repair_required_results = 0
    rejected_results = sum(1 for r in inspections if (r.overall_status or "") == "nok")
    materials_with_certificate = sum(1 for m in materials if (m.certificate_no or "").strip())
    welds_with_welder = sum(1 for w in welds if (w.welders or "").strip())
    welds_with_wps = sum(1 for w in welds if (w.wps or "").strip())
    ndt_completed = sum(1 for r in ndt_records if (r.result or "").lower() not in ("pending", "open", ""))

    completeness = [
        {
            "key": "project",
            "label": "Projectgegevens",
            "status": "complete" if (project.code and project.name) else "incomplete",
            "detail": "Projectnummer en projectnaam aanwezig." if (project.code and project.name) else "Project mist code of naam.",
        },
        {
            "key": "assemblies",
            "label": "Assemblies",
            "status": "complete" if len(assemblies) > 0 else "incomplete",
            "detail": f"{len(assemblies)} assemblies geregistreerd.",
        },
        {
            "key": "welds",
            "label": "Lassen",
            "status": "complete" if len(welds) > 0 else "incomplete",
            "detail": f"{len(welds)} lassen geregistreerd.",
        },
        {
            "key": "inspections",
            "label": "Inspecties",
            "status": "complete" if len(inspections) >= max(1, len(welds)) and len(welds) > 0 else "incomplete",
            "detail": f"{len(inspections)} inspecties voor {len(welds)} lassen.",
        },
        {
            "key": "iso5817",
            "label": "ISO 5817 beoordeling",
            "status": "complete" if len(inspections) > 0 else "incomplete",
            "detail": f"Accepted: {accepted_results}, repair required: {repair_required_results}, rejected: {rejected_results}.",
        },
        {
            "key": "ndt",
            "label": "NDT rapporten",
            "status": "complete" if len(ndt_records) > 0 else "warning",
            "detail": f"{ndt_completed}/{len(ndt_records)} NDT-rapporten met resultaat.",
        },
        {
            "key": "materials",
            "label": "Materiaalcertificaten",
            "status": "complete"
            if len(materials) > 0 and materials_with_certificate == len(materials)
            else ("warning" if materials else "incomplete"),
            "detail": f"{materials_with_certificate}/{len(materials)} materiaalrecords met certificaatnummer.",
        },
        {
            "key": "welders",
            "label": "Lassers / certificaten",
            "status": "complete"
            if len(welds) > 0 and welds_with_welder == len(welds)
            else ("warning" if welds_with_welder else "incomplete"),
            "detail": f"{welds_with_welder}/{len(welds)} lassen hebben een lasser ingevuld. {len(welder_profiles)} actieve lassers in masterdata.",
        },
        {
            "key": "wps",
            "label": "WPS / WPQR dekking",
            "status": "complete"
            if len(welds) > 0 and welds_with_wps == len(welds)
            else ("warning" if welds_with_wps else "incomplete"),
            "detail": f"{welds_with_wps}/{len(welds)} lassen hebben WPS gekoppeld. {len(wps_records)} WPS en {len(wpqr_records)} WPQR beschikbaar.",
        },
        {
            "key": "audit",
            "label": "Audit trail",
            "status": "complete" if len(audit_rows) > 0 else "warning",
            "detail": f"{len(audit_rows)} auditregels beschikbaar.",
        },
        {
            "key": "attachments",
            "label": "Bijlagen",
            "status": "complete" if project_attachment_count > 0 else "warning",
            "detail": f"{project_attachment_count} relevante bijlagen voor project/lassen/inspecties.",
        },
    ]

    ready = all(item["status"] == "complete" for item in completeness if item["key"] not in {"ndt", "attachments"})

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "project": {
            "id": _uuid(project.id),
            "code": project.code,
            "name": project.name,
            "client_name": project.client_name,
            "execution_class": project.execution_class,
            "acceptance_class": project.acceptance_class,
            "status": project.status,
            "locked": bool(project.locked),
        },
        "summary": {
            "assemblies_count": len(assemblies),
            "welds_count": len(welds),
            "inspections_count": len(inspections),
            "inspection_results_count": len(inspections),
            "defects_count": len(defects),
            "ndt_count": len(ndt_records),
            "materials_count": len(materials),
            "attachments_count": project_attachment_count,
            "audit_count": len(audit_rows),
            "accepted_results": accepted_results,
            "repair_required_results": repair_required_results,
            "rejected_results": rejected_results,
        },
        "completeness": completeness,
        "ready_for_export": ready,
        "assemblies": [
            {
                "id": _uuid(a.id),
                "code": a.code,
                "name": a.name,
                "drawing_no": a.drawing_no,
                "revision": a.revision,
                "status": a.status,
            }
            for a in assemblies
        ],
        "welds": [
            {
                "id": _uuid(w.id),
                "assembly_id": _uuid(w.assembly_id),
                "weld_no": w.weld_no,
                "location": w.location,
                "wps": w.wps,
                "process": w.process,
                "material": w.material,
                "thickness": w.thickness,
                "welders": w.welders,
                "vt_status": w.vt_status,
                "ndo_status": w.ndo_status,
                "status": w.status,
            }
            for w in welds
        ],
        "inspection_results": [
            {
                "id": _uuid(r.id),
                "weld_id": _uuid(r.weld_id),
                "status": r.overall_status,
                "inspector": r.inspector,
                "inspected_at": _dt(r.inspected_at),
                "remarks": r.remarks,
            }
            for r in inspections
        ],
        "ndt_records": [
            {
                "id": _uuid(n.id),
                "weld_id": _uuid(n.weld_id),
                "assembly_id": _uuid(n.assembly_id),
                "method": n.method,
                "inspection_date": _dt(n.inspection_date),
                "result": n.result,
                "report_no": n.report_no,
                "inspector": n.inspector,
            }
            for n in ndt_records
        ],
        "materials": [
            {
                "id": _uuid(m.id),
                "assembly_id": _uuid(m.assembly_id),
                "material_grade": m.material_grade,
                "heat_no": m.heat_no,
                "profile": m.profile,
                "dimensions": m.dimensions,
                "quantity": m.quantity,
                "certificate_no": m.certificate_no,
            }
            for m in materials
        ],
        "audit_log": [
            {
                "id": _uuid(a.id),
                "action": a.action,
                "entity": a.entity,
                "entity_id": a.entity_id,
                "created_at": _dt(a.created_at),
                "meta": a.meta,
            }
            for a in audit_rows
        ],
        "iso5817_reference_count": len(defects),
    }


def _write_json(target: Path, payload: dict[str, Any] | list[dict[str, Any]]) -> None:
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_text_summary(target: Path, preview: dict[str, Any]) -> None:
    lines = []
    project = preview["project"]
    summary = preview["summary"]
    lines.append("CE DOSSIER SAMENVATTING")
    lines.append(f"Project: {project.get('code') or '-'} - {project.get('name') or '-'}")
    lines.append(f"Opdrachtgever: {project.get('client_name') or '-'}")
    lines.append(
        f"Executieklasse: {project.get('execution_class') or '-'} | Acceptatieklasse: {project.get('acceptance_class') or '-'}"
    )
    lines.append("")
    lines.append("Tellingen")
    for key, value in summary.items():
        lines.append(f"- {key}: {value}")
    lines.append("")
    lines.append("Completeness")
    for item in preview["completeness"]:
        lines.append(f"- [{item['status']}] {item['label']}: {item['detail']}")
    target.write_text("\n".join(lines), encoding="utf-8")


def _write_pdf_summary(target: Path, preview: dict[str, Any]) -> None:
    styles = getSampleStyleSheet()
    story = []
    project = preview["project"]
    summary = preview["summary"]

    story.append(Paragraph("CE Dossier", styles["Title"]))
    story.append(Paragraph(f"Project: {project.get('code') or '-'} — {project.get('name') or '-'}", styles["Heading2"]))
    story.append(Paragraph(f"Opdrachtgever: {project.get('client_name') or '-'}", styles["Normal"]))
    story.append(
        Paragraph(
            f"EXC: {project.get('execution_class') or '-'} | Acceptatieklasse: {project.get('acceptance_class') or '-'}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))

    table_data = [["Kengetal", "Waarde"]] + [[key, str(value)] for key, value in summary.items()]
    table = Table(table_data, repeatRows=1, colWidths=[220, 220])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C4D0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph("Completeness", styles["Heading2"]))

    completeness_rows = [["Onderdeel", "Status", "Detail"]]
    for item in preview["completeness"]:
        completeness_rows.append([item["label"], item["status"], item["detail"]])

    completeness_table = Table(completeness_rows, repeatRows=1, colWidths=[140, 80, 220])
    completeness_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F75B5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(completeness_table)

    doc = SimpleDocTemplate(str(target), pagesize=A4, title="CE Dossier")
    doc.build(story)


def _write_excel_summary(target: Path, preview: dict[str, Any]) -> None:
    workbook = Workbook()

    ws_summary = workbook.active
    ws_summary.title = "Samenvatting"
    ws_summary["A1"] = "CE Dossier"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Project"
    ws_summary["B3"] = f"{preview['project'].get('code') or '-'} - {preview['project'].get('name') or '-'}"
    ws_summary["A4"] = "Opdrachtgever"
    ws_summary["B4"] = preview["project"].get("client_name") or "-"
    ws_summary["A5"] = "EXC"
    ws_summary["B5"] = preview["project"].get("execution_class") or "-"

    ws_summary["A7"] = "Kengetal"
    ws_summary["B7"] = "Waarde"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ("A7", "B7"):
        ws_summary[cell].font = Font(color="FFFFFF", bold=True)
        ws_summary[cell].fill = header_fill

    row = 8
    for key, value in preview["summary"].items():
        ws_summary[f"A{row}"] = key
        ws_summary[f"B{row}"] = value
        row += 1

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 18

    ws_check = workbook.create_sheet("Completeness")
    ws_check.append(["Onderdeel", "Status", "Detail"])
    for cell in ws_check[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="2F75B5")
    for item in preview["completeness"]:
        ws_check.append([item["label"], item["status"], item["detail"]])

    ws_check.column_dimensions["A"].width = 28
    ws_check.column_dimensions["B"].width = 16
    ws_check.column_dimensions["C"].width = 70

    ws_welds = workbook.create_sheet("Lassen")
    ws_welds.append(["Lasnummer", "Locatie", "WPS", "Proces", "Materiaal", "Dikte", "Lasser", "Status"])
    for cell in ws_welds[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for weld in preview["welds"]:
        ws_welds.append(
            [
                weld.get("weld_no"),
                weld.get("location"),
                weld.get("wps"),
                weld.get("process"),
                weld.get("material"),
                weld.get("thickness"),
                weld.get("welders"),
                weld.get("status"),
            ]
        )

    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws_welds.column_dimensions[col].width = 18

    workbook.save(target)


def generate_export_bundle(
    db: Session,
    tenant_id: UUID,
    project_id: UUID,
    *,
    user_id: UUID | None = None,
    requested_by: str | None = None,
    bundle_type: str = "zip",
) -> ExportJob:
    preview = build_preview(db, tenant_id, project_id)

    job = ExportJob(
        tenant_id=tenant_id,
        project_id=project_id,
        export_type="ce_dossier",
        status="running",
        requested_by=requested_by,
        message="CE-dossier wordt opgebouwd.",
        bundle_type=bundle_type,
        manifest_json=json.dumps({"ready_for_export": preview["ready_for_export"]}, ensure_ascii=False),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    export_dir = EXPORT_ROOT / str(project_id) / str(job.id)
    export_dir.mkdir(parents=True, exist_ok=True)

    _write_json(export_dir / "preview.json", preview)
    _write_json(export_dir / "assemblies.json", preview["assemblies"])
    _write_json(export_dir / "welds.json", preview["welds"])
    _write_json(export_dir / "inspection_results.json", preview["inspection_results"])
    _write_json(export_dir / "ndt_records.json", preview["ndt_records"])
    _write_json(export_dir / "materials.json", preview["materials"])
    _write_json(export_dir / "audit_log.json", preview["audit_log"])
    _write_text_summary(export_dir / "README_CE_DOSSIER.txt", preview)
    _write_pdf_summary(export_dir / "CE_DOSSIER_SUMMARY.pdf", preview)
    _write_excel_summary(export_dir / "CE_DOSSIER_SUMMARY.xlsx", preview)

    manifest = {
        "ready_for_export": preview["ready_for_export"],
        "generated_at": preview["generated_at"],
        "bundle_type": (bundle_type or "zip").lower(),
        "project": preview["project"],
        "files": sorted(p.name for p in export_dir.iterdir()),
    }
    _write_json(export_dir / "manifest.json", manifest)

    zip_path = export_dir / f"ce_dossier_{preview['project'].get('code') or project_id}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file in sorted(export_dir.iterdir()):
            if file == zip_path:
                continue
            zf.write(file, arcname=f"ce_dossier/{file.name}")

    final_path = zip_path
    final_message = "CE-dossier ZIP-bundel aangemaakt."
    normalized_bundle = (bundle_type or "zip").lower()
    if normalized_bundle == "pdf":
        final_path = export_dir / "CE_DOSSIER_SUMMARY.pdf"
        final_message = "CE-dossier PDF aangemaakt."
    elif normalized_bundle == "excel":
        final_path = export_dir / "CE_DOSSIER_SUMMARY.xlsx"
        final_message = "CE-dossier Excel aangemaakt."

    job.status = "completed"
    job.file_path = str(final_path)
    job.message = final_message
    job.completed_at = datetime.now(timezone.utc)
    job.manifest_json = json.dumps(
        {
            **manifest,
            "zip_name": zip_path.name,
            "download_name": final_path.name,
        },
        ensure_ascii=False,
    )
    db.commit()
    db.refresh(job)

    log_phase5_event(
        db,
        tenant_id=tenant_id,
        user_id=user_id,
        action="ce_dossier_export_created",
        entity="export_job",
        entity_id=job.id,
        meta={
            "project_id": str(project_id),
            "ready_for_export": preview["ready_for_export"],
            "bundle_type": bundle_type,
        },
    )
    return job